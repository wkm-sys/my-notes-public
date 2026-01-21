import zipfile
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime
import tkinter as tk
from tkinter import filedialog

# ===== 弹窗选择目录 =====
root = tk.Tk()
root.withdraw()  # 不显示主窗口

selected_dir = filedialog.askdirectory(title="请选择包含 ZIP 文件的目录")

if not selected_dir:
    raise SystemExit("未选择目录，程序已退出")

TARGET_DIR = Path(selected_dir)
OUTPUT_DIR = TARGET_DIR                 # log 输出位置

# ===== 数据容器（顺序 + 去重）=====
ordered_items = []
seen_items = set()

zip_count = 0
xlsx_count = 0

# ===== 扫描目录下所有 zip（按文件系统顺序）=====
for zip_path in sorted(TARGET_DIR.glob("*.zip")):
    zip_count += 1   # ← 就加在这里
    
    try:
        with zipfile.ZipFile(zip_path) as z:
            # zip 内的 xlsx 也按顺序处理
            for name in sorted(z.namelist()):
                if name.lower().endswith(".xlsx"):
                    with z.open(name) as f:
                        wb = load_workbook(f, data_only=True)
                        xlsx_count += 1   # ← 放在这里
                        
                        for ws in wb.worksheets:
                            for row in ws.iter_rows(values_only=True):
                                for cell in row:
                                    if cell is None:
                                        continue
                                    value = str(cell).strip()
                                    if not value:
                                        continue
                                    if value in seen_items:
                                        continue
                                    seen_items.add(value)
                                    ordered_items.append(value)
    except Exception as e:
        print(f"跳过异常文件: {zip_path} -> {e}")

# ===== 生成 log 文件名 =====
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
log_path = OUTPUT_DIR / f"{timestamp}.log"

# ===== 输出阶段：编号连续 + JOB 提示行 =====
with open(log_path, "w", encoding="utf-8") as f:
    index = 1
    for item in ordered_items:
        if "JOB NO" in item.upper():
            f.write("\n")
            f.write("! =========INFO===Next_JOB=========\n")
            f.write("\n")

        f.write(f"{index:04d}  {item}\n")
        index += 1

    f.write("\n")
    f.write("! INFO- ZIP COUNT: " + str(zip_count) + "\n")
    f.write("! INFO- XLSX COUNT: " + str(xlsx_count) + "\n")

     
import os

os.startfile(selected_dir)

print(f"完成，共写入 {len(ordered_items)} 条，输出文件：{log_path}")


